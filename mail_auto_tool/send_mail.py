import pandas as pd
import win32com.client as win32
from openpyxl import load_workbook
from datetime import datetime
import os

# ==============================
# 設定
# ==============================
MAIL_DATA_PATH = "mail_data.xlsx"          # 宛先リスト
TEMPLATE_PATH = "template_mail.xlsx"        # テンプレート（A1～C15を想定）
LOG_PATH = "send_log.xlsx"                  # 送信ログファイル
SEND_FLAG_COL = "SendFlag"                  # 送信フラグ列名
TEST_MODE = True                            # True: 確認表示 / False: 実際に送信

# ==============================
# メールテンプレート読込
# ==============================
def load_template(template_path):
    wb = load_workbook(template_path, data_only=True)
    ws = wb.active
    content = []
    for row in ws.iter_rows(min_row=1, max_row=15, max_col=3, values_only=True):
        line = " ".join([str(cell) for cell in row if cell is not None])
        if line.strip():
            content.append(line)
    return "\n".join(content)

# ==============================
# ログ書き込み
# ==============================
def save_log(log_list):
    log_df = pd.DataFrame(log_list, columns=["日時", "宛先", "件名", "結果"])
    
    if os.path.exists(LOG_PATH):
        old_df = pd.read_excel(LOG_PATH)
        log_df = pd.concat([old_df, log_df], ignore_index=True)
    
    log_df.to_excel(LOG_PATH, index=False)
    print(f"📝 ログを保存しました → {LOG_PATH}")

# ==============================
# メール送信処理
# ==============================
def send_mail_from_excel():
    df = pd.read_excel(MAIL_DATA_PATH)
    outlook = win32.Dispatch("Outlook.Application")
    template_body = load_template(TEMPLATE_PATH)
    logs = []

    for _, row in df.iterrows():
        if str(row.get(SEND_FLAG_COL, "")) == "1":
            mail = outlook.CreateItem(0)
            mail.To = row["To"]
            mail.Subject = row["Subject"]

            if "{body}" in template_body:
                body_text = template_body.replace("{body}", str(row.get("Body", "")))
            else:
                body_text = template_body + "\n" + str(row.get("Body", ""))

            mail.Body = body_text

            try:
                if TEST_MODE:
                    mail.Display()  # 送信前に内容を確認
                    result = "プレビュー表示"
                else:
                    mail.Send()
                    result = "送信完了"
                print(f"✅ {result}: {row['To']}")
            except Exception as e:
                print(f"❌ 送信失敗: {row['To']} - {e}")
                result = f"送信失敗 ({e})"

            logs.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"), row["To"], row["Subject"], result])

    # ログ出力
    save_log(logs)
    print("🎉 処理が完了しました。")

# ==============================
# メイン実行
# ==============================
if __name__ == "__main__":
    send_mail_from_excel()
